#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
import os.path,sys

class excel_data():
	def __init__(self,ticker):
		self.wb = Workbook()
		self.worksheet = 0
		self.ticker = ticker.upper()
		reload(sys)
		sys.setdefaultencoding('UTF8')

	def open_sheet(self):
		if not (os.path.isfile('/home/vivek/Desktop/Stefano/Code/Algo1/Data/'+self.ticker+'.xlsx')):
			self.wb = Workbook()
			self.worksheet = self.wb.create_sheet(0)
			self.worksheet.title = self.ticker+" NEWS"
			tags = ['Date','Time','Headline','Description','Link']
			self.worksheet.append(tags)
		else:
			self.wb = load_workbook('/home/vivek/Desktop/Stefano/Code/Algo1/Data/'+self.ticker+'.xlsx')
			self.worksheet = self.wb[self.ticker+" NEWS"]
 
	def write_sheet(self,data):
		chk_count =  int(self.worksheet.get_highest_row())
		if (chk_count -60) <=0:
			chk_start = 1
		else :
			chk_start = chk_count-60
		wrt_token = True
		for row in self.worksheet.iter_rows('C'+str(chk_start)+':C'+str(chk_count)+''):
			for cell in row:
				if data[2] == cell.value:
					wrt_token = False 
					break

		if (wrt_token):
			print "**** "+data[2]+" ****"
			self.worksheet.append(data)
			return data[2]
		else :
			return False

	def reading_sheet(self):
		chk_count =  int(self.worksheet.get_highest_row())
		output_data = []
		for i in range(2,chk_count+1):
			output_data.append(self.worksheet['C'+str(i)+''].value)
		return output_data

	def close_sheet(self):
		self.wb.save('/home/vivek/Desktop/Stefano/Code/Algo1/Data/'+self.ticker+'.xlsx')


if __name__=="__main__":
	obj = excel_data()
	data = ['07-06-2016','9:12:36','This is amazing','All of it is amazing','htttp://visit.me']
	obj.write_sheet(data)