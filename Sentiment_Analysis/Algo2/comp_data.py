from openpyxl import Workbook,load_workbook
import os, sys, time, itertools
from selenium import webdriver

class comp_goog_data():
    def __init__(self,ticker):
        chromedriver = "/home/vivek/Desktop/Stefano/Code/Algo2/chromedriver"
        os.environ["webdriver.chrome.driver"] = chromedriver
        self.driver = webdriver.Chrome(chromedriver)
        self.driver.get("https://www.google.com/#newwindow=1&q="+ticker+'%20ticker')
        time.sleep(5)

    def main_page(self,ticker):
        self.driver.get("https://www.google.com/#newwindow=1&q="+ticker+'%20ticker')
        time.sleep(5)

    def get_name(self):
        return [str(self.xpath_check('//*[@id="rhs_block"]/div/div[1]/div/div[1]/div[2]/div[1]/div/div[2]/div/div/div[2]/div[1]','text'))]

    def get_comp_sector(self):
        return [str(self.xpath_check('//*[@id="rhs_block"]/div/div[1]/div/div[1]/div[2]/div[1]/div/div[2]/div/div/div[2]/div[2]/span','text'))]

    def get_ticker(self):
        name = self.get_name()
        if name :
            comp_name = name[0].replace(' ', '%20')
            self.driver.get("https://www.google.com/#newwindow=1&q="+comp_name+"%20ticker")
            time.sleep(5)
            ticker = self.xpath_check('//*[@id="fac-ut"]/div[1]/div[2]/span[1]','text')
            if ticker: 
                ticker = str(ticker.split(':')[1])
            else :
                ticker = ' No ticker'
        else :
            ticker = ' No ticker'
        return [ticker[1:]]

    def get_subs(self):
        fields = self.get_info()
        subs = []
        for key,value in fields.items():
            if key == 'Subsidiaries':
                self.xpath_check(value,'click')
                time.sleep(5)
                subs = self.get_sub_list()
                if not subs:
                    subs = self.get_comp_list()
        return subs

    def get_parents(self):
        fields = self.get_info()
        parents= []
        for key,value in fields.items():
            if key == 'Parent organizations':
                self.xpath_check(value,'click')
                time.sleep(5)
                parents = self.get_sub_list()
                if not parents:
                    parents = self.get_comp_list()
        return parents

    def get_ceo_name(self):
        fields = self.get_info()
        ceo = ''
        for key,value in fields.items():
            if key == 'CEO':
                value = value.replace('span[1]/a','span[2]/a')
                ceo = self.xpath_check(value,'text')
        return [str(ceo)]

    def competitors(self):
        if not xpath_check('//*[@id="rhs_block"]/div/div[1]/div/div[1]/div[2]/div[4]/div[1]/a','click'): #competitors
            competitors = get_comp_list()
        return competitors

    def xpath_check(self,xpath,action):
        time.sleep(1)
        if action == 'text':
            try : 
               value = self.driver.find_element_by_xpath(xpath).text
            except Exception as e:
                value = False
                #print xpath,' : xpath not good : ', e
        elif action == 'click':
            try : 
                value = self.driver.find_element_by_xpath(xpath).click()
            except Exception as e:
                value = False
                #print xpath,' : xpath not good : ', e
        return value


    def get_sub_list(self):
        comp_list,loop,row,col,empty = [],True,1,1,False
        # loop = self.xpath_check('//*[@id="uid_0"]/div/div/div[1]/div/a/div[2]/div/div/div','text')
        # if not loop :
        #     print 'exit'
        #     loop = self.xpath_check('//*[@id="uid_0"]/div/div/div[1]/div/a/div[2]/div[1]/div/div','text')
        # if loop : 
        #     print 'enter:: ',loop
        #     comp_list.append(str(loop))
        # while True:  
        #     print 'first'
        #     text = self.xpath_check('//*[@id="uid_0"]/div/div/div[1]/div/a['+str(row)+']/div[2]/div/div/div','text')
        #     if not text:
        #         text = self.xpath_check('//*[@id="uid_0"]/div/div/div[1]/div/a['+str(row)+']/div[2]/div[1]/div/div','text')
        #     if not text:  
        #         break     
        #     comp_list.append(str(text))
        #     row +=1       
        while True:         
            # print 'second'  
            # loop = self.xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/div/a/div[2]/div/div/div','text')
            # if not loop :
            #     loop = self.xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/div/a/div[2]/div[1]/div/div','text')
            # if loop :
            #     print 'loop :: ',loop
            #     comp_list.append(str(loop))
            row=1
            while True:
                text = self.xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/div/a['+str(row)+']/div[2]/div/div/div','text')
                if not text:
                    text = self.xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/div/a['+str(row)+']/div[2]/div[1]/div/div','text')
                if not text:
                    if row == 1:
                        empty = True
                        break
                    else :
                        break
                comp_list.append(str(text))
                row +=1
            if empty :
                break
            col+=1
        return comp_list

    def get_comp_list(self):
        comp_list,col = [],1
        while True:
            text = self.xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/a/div[2]/div[1]/span','text')
            if not text:
                break
            comp_list.append(str(text)) #competitors list #comp_lisself.t.append(xpath_check('//*[@id="uid_0"]/div/div/div['+str(col)+']/a','click')) #competitors
            col +=1
        return comp_list

    def get_info(self):
        fields = {}
        x_path_1 = '//*[@id="rhs_block"]/div/div[1]/div/div[1]/div[2]/div['
        x_path_2 = ']/div/div['
        x_path_3 = ']/div/div/span[1]/a'
        for i in range(2,5):
            for j in range(1,10):
               fields[str(self.xpath_check(x_path_1+str(i)+x_path_2+str(j)+x_path_3,'text'))] = x_path_1 +str(i)+x_path_2+str(j)+x_path_3
        fields.pop('False', None)
        return fields

    def ticker_comp_details(self):
        comp_data = []
        name =  self.get_name()
        ticker = self.get_ticker()
        subs = self.get_subs()
        parent = self.get_parents()
        ceo = self.get_ceo_name()
        comp_data.append([name,ticker,subs,parent,ceo])
        if parent:
            for i in parent:
                self.main_page(i)
                subs_p = self.get_subs()
                parent_p = self.get_parents()
                ceo_p = self.get_ceo_name()
                comp_data.append([subs_p,parent_p,ceo_p])
        return list(itertools.chain.from_iterable(list(itertools.chain.from_iterable(comp_data))))

    def __del__(self):
        self.driver.quit()


class comp_file_data():
	def __init__(self,ticker):
		self.wb = Workbook()
		self.worksheet = 0
		self.ticker = str(ticker.upper())
		reload(sys)
		sys.setdefaultencoding('UTF8')
		self.con_lst = ['NorthAmerica','Europe','Asia']

	def open_sheet(self,file):
		self.wb = load_workbook('/home/vivek/Desktop/Stefano/Code/Algo2/company_data/continent_list/'+file+'.xlsx')
		self.worksheet = self.wb[file]

	def read_sheet(self):
		for file in range(len(self.con_lst)):
			self.open_sheet(str(self.con_lst[file]))
			chk_count =  int(self.worksheet.get_highest_row())
			for i in range(2,chk_count+1):
				if str(self.worksheet['A'+str(i)+''].value) == self.ticker:
					output_data = {'Symbol':self.ticker ,'Name':str(self.worksheet['B'+str(i)+''].value) ,'Country':str(self.worksheet['F'+str(i)+''].value) ,'Sector':str(self.worksheet['H'+str(i)+''].value) ,'Industry':str(self.worksheet['I'+str(i)+''].value) }
					return output_data

if __name__=="__main__":
    data = comp_goog_data('Google')
    print data.ticker_comp_details()
